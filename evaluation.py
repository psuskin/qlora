import os
import openpyxl
from collections import Counter
from collections import OrderedDict
from inference import load_model, generate

RESET = False

filename = "evaluation.xlsx"

prompts = OrderedDict([
    ("general", [
        "What is a module?",

        "In what module can I edit customers?",
        "In what module do I edit the name of a customer?",

        "What is the module for entering sales invoices?",
    ]),
    ("hallucination", [
        "This is the context of the module billofma: feeding guinea pigs and groundhogs. Which module describes feeding mammals?",
        "This is the description of the module billofma: Feeding guinea pigs and groundhogs. Which module describes Donald Trump's presidency?",
    ]),
    ("specific", {
        "balanfac": [
            "With this module, the annual and period balances of a general ledger or personal account posted in financial accounting are displayed. Which module is being described?",
            "Which module is used to display the annual and period balances of a general ledger?",
            "What is the module to list annual balances of general ledger?",
        ],
        "icastedt": [
            "Which module deals with creating and deleting parts or service-role relationships?",
        ],
        "billofma": [
            "Which module describes the composition of a production part?"
        ],
    }),
    ("realistic (finding module)", [
        "Wo erfasse ich Kunden?",
        "Wo gebe ich Arbeitsplatz Kapazitäten ein?",
        "Wo erfasse ich den Urlaub eines Mitarbeiters?",
        "Wie kann ich die Gültigkeit eines Teils beschränken?",
        "In welchem Modul gibt man das Logistik Kennzeichen ein?",
        "Wie erfasse ich eine Rahmenbestellung?",
        "Wo pflege ich Währungstabellen?",
    ]),
    ("realistic (explanation)", [
        "Was ist ein Variantenteil?",
        "Was macht der Gozintograph?",
        "Wie plane ich Fertigungsaufträge?",
        "Was sind Spezifikationsnummern?",
        "Was ist eine Sachmerkmalsleiste?",
        "Wie setze ich Attribute in die Sachmerkmalsleiste ein?",
        "Wie werden die Teile im Lager bewertet?",
        "Was ist eine Preistabelle?",
        "Was sind bedingte Stücklistenpositionen?",
        "Was ist eine Fehlzeit?",
        "Wie importiere ich meine Inventurdaten in GESTIN?",
        "Was ist eine Stichprobeninventur?"
        "Was macht PYTHIA?",
        "Kann man Belege löschen?",
        "Was ist die Beleghistorie?",
        "Kann ich die Ausgabewährung einer Auftragsbestätigung anpassen?",
        "Was ist ein Logistik-Kennzeichen?",
        "Was ist eine Beipackliste?",
        "Was sind Stammdaten?",
    ])
])

models =  OrderedDict([
    ("LLaMA-7b", {
        "path": "huggyllama/llama-7b",
        "motivation": "Able to run on local PC, reputable foundation model."
    }),
    ("LLaMA-2-7b", {
        "path": "meta-llama/Llama-2-7b-hf",
        "motivation": "Fill out model evaluation."
    }),
    ("LLaMA-2-7b-chat", {
        "path": "meta-llama/Llama-2-7b-chat-hf",
        "motivation": "Check if finetuning an instruction-finetuned model improves chat dialogue."
    }),
    ("LLaMA-2-13b", {
        "path": "meta-llama/Llama-2-13b-hf",
        "motivation": "Fill out model evaluation."
    }),
    ("LLaMA-2-13b-chat", {
        "path": "meta-llama/Llama-2-13b-chat-hf",
        "motivation": "Fill out model evaluation."
    }),
    ("LLaMA-2-70b", {
        "path": "meta-llama/Llama-2-70b-hf",
        "motivation": "Acquire a feel for the runtime and performance associated with a large model."
    }),
])

datasets = OrderedDict([
    # ("I/O falsch", {
    #     "format": [
    #         "[], [###Context###: This is the description of the module <Modulename>: <Module Description>.\n###Instruction###: What is the name of this module?\n###Response###: The name of the module is <Modulename>.]",
    #         "[], [###Context###:\n###Instruction###: What is the purpose of the module <Modulename>?\n###Response###: The purpose of the module <Modulename> is <Module Description>.]"
    #     ],
    #     "motivation": "The main goal is for the model to answer correct module names based on descriptions, supported by the first data format. The second data format is to give the model an understanding of appsWH."
    # }),
    # ("Alpaca", {
    #     "format": [
    #         "[Below is an instruction that describes a task, paired with an input that provides further context. Write a response that appropriately completes the request.\n\n### Instruction:\nWhat is the name of this module?\n\n### Context:\nThis is the description of the module <Modulename>: <Module Description>\n\n### Response:], [The name of the module is <Modulename>.]",
    #         "[Below is an instruction that describes a task. Write a response that appropriately completes the request.\n\n### Instruction:\nWhat is the purpose of the module <Modulename>?\n\n### Response:], [The purpose of the module <Modulename> is <Module Description>.]"
    #     ],
    #     "motivation": "The previous data format put all data in the output, causing the model to recursively generate Context and Instruction tags. This data format is the same as the previous one, but only the response is in the output."
    # }),
    ("Alpaca ohne Modulnamen im Kontext", {
        "format": [
            "[Below is an instruction that describes a task, paired with an input that provides further context. Write a response that appropriately completes the request.\n\n### Instruction:\nWhat is the name of this module?\n\n### Context:\n<Module Description>\n\n### Response:], [The name of the module is <Modulename>.]",
            "[Below is an instruction that describes a task. Write a response that appropriately completes the request.\n\n### Instruction:\nWhat is the purpose of the module <Modulename>?\n\n### Response:], [The purpose of the module <Modulename> is <Module Description>.]"
        ],
        "motivation": "The previous data format gave away the module name in the context, causing the model to pay attention only to the passed name instead of the actual context. This data format is the same as the previous one, but the module name is removed from the context."
    }),
    # ("Alpaca refined", {
    #     "format": [
    #         "[Below is an instruction that describes a task, paired with an input that provides further context. Write a response that appropriately completes the request.\n\n### Instruction:\nWhat is the name of this module?\n\n### Input:\n<Module Description>\n\n### Response:], [The name of the module is <Modulename>.]",
    #         "[Below is an instruction that describes a task. Write a response that appropriately completes the request.\n\n### Instruction:\nWhat is the purpose of the module <Modulename>?\n\n### Response:], [The purpose of the module <Modulename> is <Module Description>.]"
    #     ],
    #     "motivation": "The previous data format used the Context tag. In order to finetune the already instruction tuned LLaMA-2-chat model, the Input tag is used instead. This data format is the same as the previous one, but the Context tag is replaced with Input."
    # }),
    ("Alpaca ohne Modulnamen im Kontext (de)", {
        "format": [
            "[Unten ist eine Anweisung, die eine Aufgabe beschreibt, gepaart mit einer Eingabe, die weitere Kontextinformationen liefert. Schreiben Sie eine Antwort, die die Anfrage angemessen vervollständigt.\n\n### Anweisung:\nWie heißt dieses Modul?\n\n### Kontext:\n<Module Description>\n\n### Antwort:], [Der Name des Moduls ist <Modulename>.]",
            "[Unten ist eine Anweisung, die eine Aufgabe beschreibt. Schreiben Sie eine Antwort, die die Anfrage angemessen vervollständigt.\n\n### Anweisung:\nWas ist der Zweck des Moduls <Modulename>?\n\n### Antwort:], [Der Zweck des Moduls <Modulename> ist <Module Description>.]"
        ],
        "motivation": "Test LLaMA-7b's multilingualism."
    }),
    ("Alpaca ohne Modulnamen im Kontext (extratokens)", {
        "format": [
            "[Below is an instruction that describes a task, paired with an input that provides further context. Write a response that appropriately completes the request.\n\n### Instruction:\nWhat is the name of this module?\n\n### Context:\n<Module Description>\n\n### Response:], [The name of the module is <Modulename>.]",
            "[Below is an instruction that describes a task. Write a response that appropriately completes the request.\n\n### Instruction:\nWhat is the purpose of the module <Modulename>?\n\n### Response:], [The purpose of the module <Modulename> is <Module Description>.]"
        ],
        "motivation": "See if adding module names to the token vocabulary produces more accurate module names."
    }),
])

adapters = OrderedDict([
    # ("guanaco-7b", {
    #     "model": "LLaMA-7b",
    #     "dataset": "I/O falsch",
    #     "checkpoint": "1875"
    # }),
    # ("alpaca-7b", {
    #     "model": "LLaMA-7b",
    #     "dataset": "Alpaca",
    #     "checkpoint": "1875"
    # }),
    ("alpaca-7b-en", {
        "model": "LLaMA-7b",
        "dataset": "Alpaca ohne Modulnamen im Kontext",
        "checkpoint": "1875"
    }),
    ("alpaca-2-7b", {
        "model": "LLaMA-2-7b",
        "dataset": "Alpaca ohne Modulnamen im Kontext",
        "checkpoint": "1875"
    }),
    ("chat-7b", {
        "model": "LLaMA-2-7b-chat",
        "dataset": "Alpaca ohne Modulnamen im Kontext",#"Alpaca refined",
        "checkpoint": "1875"
    }),
    ("alpaca-2-13b", {
        "model": "LLaMA-2-13b",
        "dataset": "Alpaca ohne Modulnamen im Kontext",
        "checkpoint": "1875"
    }),
    ("chat-13b-context", {
        "model": "LLaMA-2-13b-chat",
        "dataset": "Alpaca ohne Modulnamen im Kontext",
        "checkpoint": "1875"
    }),
    ("alpaca-2-70b", {
        "model": "LLaMA-2-70b",
        "dataset": "Alpaca ohne Modulnamen im Kontext",
        "checkpoint": "1000"
    }),
    ("alpaca-7b-de", {
        "model": "LLaMA-7b",
        "dataset": "Alpaca ohne Modulnamen im Kontext (de)",
        "checkpoint": "1875"
    }),
    ("alpaca-7b-extratokens", {
        "model": "LLaMA-7b",
        "dataset": "Alpaca ohne Modulnamen im Kontext (extratokens)",
        "checkpoint": "1875"
    }),
])

def readInferences():
    inferences = {}
    if os.path.isfile(filename):
        wb = openpyxl.load_workbook(filename)
        ws = wb["Prompts"]

        currentAdapters = []
        for col in ws.iter_cols(min_col=2, max_row=1, max_col=ws.max_column):
            for cell in col:
                currentAdapters.append(cell.value)
                inferences[cell.value] = {}
        
        for row in ws.iter_rows(min_row=2, max_col=ws.max_column, max_row=ws.max_row):
            for j, cell in enumerate(row):
                if j == 0:
                    if cell.font.bold:
                        break
                    else:
                        prompt = cell.value
                else:
                    if cell.value:
                        inferences[currentAdapters[j-1]][prompt] = cell.value.strip()

    return inferences

def infer():
    inferences = readInferences()

    if False:
        inferences = {}
        adapterNames = list(adapters.keys())
        for name in adapterNames:
            inferences[name] = {}

            for promptCategory in list(prompts.keys()):
                if promptCategory == "specific":
                    for module in prompts[promptCategory]:
                        for prompt in prompts[promptCategory][module]:
                            inferences[name][prompt] = "test"
                else:
                    for prompt in prompts[promptCategory]:
                        inferences[name][prompt] = "test"

        return inferences

    currentPrompts = []
    for promptCategory in list(prompts.keys()):
        if promptCategory == "specific":
            for module in prompts[promptCategory]:
                for prompt in prompts[promptCategory][module]:
                    currentPrompts.append(prompt)
        else:
            for prompt in prompts[promptCategory]:
                currentPrompts.append(prompt)

    adapterNames = list(adapters.keys())
    for name in adapterNames:
        if name in inferences and Counter(list(inferences[name].keys())) == Counter(currentPrompts) and not RESET:
            continue

        base_model, finetuned_model, tokenizer = load_model(models[adapters[name]["model"]]["path"], f"/workspace/output/{name}/checkpoint-{adapters[name]['checkpoint']}/adapter_model")

        for promptCategory in list(prompts.keys()):
            if promptCategory == "specific":
                for module in prompts[promptCategory]:
                    for prompt in prompts[promptCategory][module]:
                        if prompt in inferences[name] and not RESET:
                            continue

                        try:
                            inferences[name][prompt] = generate(finetuned_model, tokenizer, prompt, True, not "de" in name) + generate(base_model, tokenizer, prompt, True, not "de" in name)
                        except:
                            inferences[name][prompt] = "Error"
            else:
                for prompt in prompts[promptCategory]:
                    if prompt in inferences[name] and not RESET:
                        continue

                    try:
                        inferences[name][prompt] = generate(finetuned_model, tokenizer, prompt, True, not "de" in name) + generate(base_model, tokenizer, prompt, True, not "de" in name)
                    except:
                        inferences[name][prompt] = "Error"

    return inferences

colors = [
    "dd776e",
    "e0816d",
    "e2886c",
    "e5926b",
    "e79a69",
    "e9a268",
    "ecac67",
    "e6ad61",
    "e9b861",
    "f3c563",
    "f5ce62",
    "e2c965",
    "d4c86a",
    "c4c56d",
    "b0be6e",
    "a4c073",
    "94bd77",
    "84bb7b",
    "73b87e",
    "63b682",
    "57bb8a"
]

def accuracy(inferences, name):
    correct = 0
    total = 0
    for module in prompts["specific"]:
        for prompt in prompts["specific"][module]:
            correct += module in inferences[name][prompt]
            total += 1

    return colors[int((correct / total) * len(colors))]

def evaluate(ws, inferences):
    adapterNames = list(adapters.keys())
    modelNames = list(models.keys())
    datasetNames = list(datasets.keys())

    datasetHeaderCell = ws["A1"]
    datasetHeaderCell.alignment = openpyxl.styles.Alignment(horizontal="center")
    datasetHeaderCell.value = "Dataset"
    modelHeaderCell = ws["B1"]
    modelHeaderCell.alignment = openpyxl.styles.Alignment(horizontal="center")
    modelHeaderCell.value = "Model"
    ws.merge_cells(start_row=1, end_row=1, start_column=2, end_column=len(modelNames)+1)

    for col in ws.iter_cols(min_row = 2, max_row=2, min_col=2, max_col=len(modelNames)+1):
        for cell in col:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.border = openpyxl.styles.borders.Border(bottom=openpyxl.styles.borders.Side(style="thin"))
            cell.value = modelNames[cell.column - 2]

    for row in ws.iter_rows(max_col=1, min_row=3, max_row=len(datasetNames)+2):
        for cell in row:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.border = openpyxl.styles.borders.Border(right=openpyxl.styles.borders.Side(style="thin"))
            cell.value = datasetNames[cell.row - 3]

    for name in adapterNames:
        c = modelNames.index(adapters[name]["model"]) + 2
        r = datasetNames.index(adapters[name]["dataset"]) + 3
        cell = ws.cell(row=r, column=c)
        cell.value = name
        color = accuracy(inferences, name)
        cell.fill = openpyxl.styles.PatternFill(start_color=color, end_color=color, fill_type="solid")

    dim_holder = openpyxl.worksheet.dimensions.DimensionHolder(worksheet=ws)
    dim_holder["A"] = openpyxl.worksheet.dimensions.ColumnDimension(ws, min=1, max=1, width=50)
    for col in range(ws.min_column+1, ws.max_column + 1):
        dim_holder[openpyxl.utils.get_column_letter(col)] = openpyxl.worksheet.dimensions.ColumnDimension(ws, min=col, max=col, width=20)
    ws.column_dimensions = dim_holder

def prompt(ws, inferences):
    adapterNames = list(adapters.keys())

    i = 1
    for promptCategory in list(prompts.keys()):
        cell = ws[f"A{i}"]
        cell.font = openpyxl.styles.Font(bold=True)
        cell.value = promptCategory.capitalize()
        i += 1
        if promptCategory == "specific":
            for module in prompts[promptCategory]:
                cell = ws[f"A{i}"]
                cell.font = openpyxl.styles.Font(bold=True)
                cell.value = module
                i += 1
                for prompt in prompts[promptCategory][module]:
                    cell = ws[f"A{i}"]
                    cell.alignment = openpyxl.styles.Alignment(wrapText=True, vertical="top")
                    cell.value = prompt
                    i += 1
        else:
            for prompt in prompts[promptCategory]:
                cell = ws[f"A{i}"]
                cell.alignment = openpyxl.styles.Alignment(wrapText=True, vertical="top")
                cell.value = prompt
                i += 1

    for a, name in enumerate(adapterNames):
        i = 1

        cell = ws[f"{openpyxl.utils.get_column_letter(a + 2)}{i}"]
        cell.font = openpyxl.styles.Font(bold=True)
        cell.value = name

        for promptCategory in list(prompts.keys()):
            i += 1
            if promptCategory == "specific":
                for module in prompts[promptCategory]:
                    i += 1
                    for prompt in prompts[promptCategory][module]:
                        cell = ws[f"{openpyxl.utils.get_column_letter(a + 2)}{i}"]
                        cell.alignment = openpyxl.styles.Alignment(wrapText=True, vertical="top")
                        cell.value = inferences[name][prompt]
                        i += 1
            else:
                for prompt in prompts[promptCategory]:
                    cell = ws[f"{openpyxl.utils.get_column_letter(a + 2)}{i}"]
                    cell.alignment = openpyxl.styles.Alignment(wrapText=True, vertical="top")
                    cell.value = inferences[name][prompt]
                    i += 1

    dim_holder = openpyxl.worksheet.dimensions.DimensionHolder(worksheet=ws)
    for col in range(ws.min_column, ws.max_column + 1):
        dim_holder[openpyxl.utils.get_column_letter(col)] = openpyxl.worksheet.dimensions.ColumnDimension(ws, min=col, max=col, width=80)
    ws.column_dimensions = dim_holder

def dataset(ws):
    headers = ["Name", "Format", "Motivation"]

    for col in ws.iter_cols(max_row=1, max_col=len(headers)):
        for cell in col:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.value = headers[cell.column - 1]

    for row in ws.iter_rows(min_row=2, max_row=len(datasets)+1, max_col=3):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrapText=True, vertical="top")
            if cell.column == 1: # name
                cell.value = list(datasets.items())[cell.row - 2][0]
            else:
                text = list(datasets.items())[cell.row - 2][1][headers[cell.column - 1].lower()]
                if isinstance(text, list):
                    text = "\n".join([f"{i+1}. {t}" for i, t in enumerate(text)])

                cell.value = text

    dim_holder = openpyxl.worksheet.dimensions.DimensionHolder(worksheet=ws)
    dim_holder[openpyxl.utils.get_column_letter(headers.index("Name") + 1)] = openpyxl.worksheet.dimensions.ColumnDimension(ws, min=headers.index("Name") + 1, max=headers.index("Name") + 1, width=30)
    dim_holder[openpyxl.utils.get_column_letter(headers.index("Format") + 1)] = openpyxl.worksheet.dimensions.ColumnDimension(ws, min=headers.index("Format") + 1, max=headers.index("Format") + 1, width=100)
    dim_holder[openpyxl.utils.get_column_letter(headers.index("Motivation") + 1)] = openpyxl.worksheet.dimensions.ColumnDimension(ws, min=headers.index("Motivation") + 1, max=headers.index("Motivation") + 1, width=80)
    ws.column_dimensions = dim_holder

def model(ws):
    headers = ["Name", "Path", "Motivation"]

    for col in ws.iter_cols(max_row=1, max_col=len(headers)):
        for cell in col:
            cell.font = openpyxl.styles.Font(bold=True)
            cell.value = headers[cell.column - 1]

    for row in ws.iter_rows(min_row=2, max_row=len(models)+1, max_col=3):
        for cell in row:
            cell.alignment = openpyxl.styles.Alignment(wrapText=True, vertical="top")
            if cell.column == 1: # name
                cell.value = list(models.items())[cell.row - 2][0]
            else:
                text = list(models.items())[cell.row - 2][1][headers[cell.column - 1].lower()]
                if isinstance(text, list):
                    text = "\n".join([f"{i+1}. {t}" for i, t in enumerate(text)])

                cell.value = text

    dim_holder = openpyxl.worksheet.dimensions.DimensionHolder(worksheet=ws)
    dim_holder[openpyxl.utils.get_column_letter(headers.index("Name") + 1)] = openpyxl.worksheet.dimensions.ColumnDimension(ws, min=headers.index("Name") + 1, max=headers.index("Name") + 1, width=20)
    dim_holder[openpyxl.utils.get_column_letter(headers.index("Path") + 1)] = openpyxl.worksheet.dimensions.ColumnDimension(ws, min=headers.index("Path") + 1, max=headers.index("Path") + 1, width=30)
    dim_holder[openpyxl.utils.get_column_letter(headers.index("Motivation") + 1)] = openpyxl.worksheet.dimensions.ColumnDimension(ws, min=headers.index("Motivation") + 1, max=headers.index("Motivation") + 1, width=80)
    ws.column_dimensions = dim_holder

if __name__ == "__main__":
    if False:
        print(readInferences())
        exit()

    wb = openpyxl.Workbook()
    wsEval = wb.active
    wsEval.title = "Evaluation"
    wsPrompts = wb.create_sheet("Prompts")
    wsDatasets = wb.create_sheet("Datasets")
    wsModels = wb.create_sheet("Models")

    inferences = infer()
    #print(inferences)

    evaluate(wsEval, inferences)

    prompt(wsPrompts, inferences)

    dataset(wsDatasets)

    model(wsModels)

    wb.save(filename)